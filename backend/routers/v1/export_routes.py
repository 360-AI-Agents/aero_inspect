import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill

from backend.dependencies import get_db, require_admin
from backend.models.user import User
from backend.models.worker import Worker
from backend.models.worker_violation import WorkerViolation
from backend.models.inspection import Inspection
from backend.models.violation import Violation

router = APIRouter(prefix="/export", tags=["Export"])

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="059669", end_color="059669", fill_type="solid")


def _style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _autosize_columns(ws, num_cols):
    for col in range(1, num_cols + 1):
        letter = ws.cell(row=1, column=col).column_letter
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 45)


def _to_excel_response(wb, filename):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/users")
def export_users(db: Session = Depends(get_db), admin=Depends(require_admin)):
    users = db.query(User).order_by(User.username).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    headers = ["Username", "Email", "Role", "Status", "Created At"]
    ws.append(headers)

    for u in users:
        ws.append([
            u.username,
            u.email,
            u.role,
            "Active" if u.is_active else "Disabled",
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
        ])

    _style_header(ws, len(headers))
    _autosize_columns(ws, len(headers))
    return _to_excel_response(wb, "aeroinspect_users.xlsx")


@router.get("/workers")
def export_workers(db: Session = Depends(get_db), admin=Depends(require_admin)):
    workers = db.query(Worker).order_by(Worker.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Workers"
    headers = ["Employee ID", "Name", "Phone", "Company", "Role", "Assigned Site",
               "Shift Start", "Shift End", "Weekly Off", "Linked Tracking ID", "Registered On"]
    ws.append(headers)

    for w in workers:
        ws.append([
            w.employee_id,
            w.name,
            w.phone or "",
            w.company or "",
            w.role or "",
            w.assigned_site,
            w.shift_start or "",
            w.shift_end or "",
            w.weekly_off_day or "",
            w.tracked_worker_id if w.tracked_worker_id is not None else "Not linked",
            w.created_at.strftime("%Y-%m-%d %H:%M") if w.created_at else "",
        ])

    _style_header(ws, len(headers))
    _autosize_columns(ws, len(headers))
    return _to_excel_response(wb, "aeroinspect_workers.xlsx")


@router.get("/violations")
def export_violations(db: Session = Depends(get_db), admin=Depends(require_admin)):
    records = db.query(WorkerViolation).order_by(WorkerViolation.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worker Violations"
    headers = ["Worker (Tracking ID)", "Linked Employee ID", "Violations", "Status",
               "First Seen", "Last Seen", "Duration (min)", "Repeat Offender", "Inspection ID"]
    ws.append(headers)

    for r in records:
        duration_min = round((r.duration_seconds or 0) / 60, 1)
        ws.append([
            r.worker_id,
            r.employee_id or "Not linked",
            r.violations or "",
            r.status,
            r.first_seen.strftime("%Y-%m-%d %H:%M") if r.first_seen else "",
            r.last_seen.strftime("%Y-%m-%d %H:%M") if r.last_seen else "",
            duration_min,
            "Yes" if r.repeat_offender else "No",
            r.inspection_id,
        ])

    _style_header(ws, len(headers))
    _autosize_columns(ws, len(headers))
    return _to_excel_response(wb, "aeroinspect_violations.xlsx")


@router.get("/reports")
def export_reports(db: Session = Depends(get_db), admin=Depends(require_admin)):
    inspections = db.query(Inspection).order_by(Inspection.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspection Reports"
    headers = ["Inspection Name", "Camera ID", "Workers Detected", "Total Violations",
               "Compliance Score (%)", "Status", "Created At"]
    ws.append(headers)

    for insp in inspections:
        ws.append([
            insp.inspection_name,
            insp.camera_id if insp.camera_id is not None else "—",
            insp.workers_detected,
            insp.total_violations,
            insp.compliance_score,
            insp.inspection_status,
            insp.created_at.strftime("%Y-%m-%d %H:%M") if insp.created_at else "",
        ])

    _style_header(ws, len(headers))
    _autosize_columns(ws, len(headers))
    return _to_excel_response(wb, "aeroinspect_reports.xlsx")